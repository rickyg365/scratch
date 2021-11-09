import os

from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


"""
OPENPYXL Notes:

# Load Workbook
filepath = "./workbooks/grades.xlsx"
wb = load_workbook(filepath)

# Load Sheet
ws = wb.active  # Active Sheet

sheet_names = wb.sheetnames  # List of sheet names

sheet_name = "Sheet1"
sheet = wb[sheet_name]  # Select Sheet

# Create Sheet
wb.create_sheet("Test")  # Does not recreate if already created

# Load Cell
cell = ws['A1']

# Load Value
cell_value = cell.value

# Change Value
new_value = "yolo"

ws["A2"] = new_value
ws["A3"].value = new_value

# Save Workbook
new_filepath = "./workbooks/pygrades.xlsx"
wb.save(new_filepath)


# Create new workbook
new_wb = Workbook()  # Overwrites if you save over existing workbook
new_ws = wb.active

# Change sheet title
new_ws.title = "NewPage"

data = ["yolo1", "yolo2", "yolo3", "yolo4", "yolo5"]
new_ws.append(data)  # Adds to first row
# append keeps adding to the next row everytime its called

# loop through
for row in range(1, 11):
    for col in range(1,5):
        # 65 == A, range(4)
        # char = chr(65 + col) or
        char = get_column_letter(col)
        print(ws[f"{char}{row}"])  # Prints cell obj
        print(ws[f"{char}{row}"].value)  # Prints cell value

new_wb.save("yolo.xlsx")


# Merge Cells
ws.merge_cells("A1:D1")  # Keeps data from the first cell

# Unmerge Cells
ws.unmerge_cells("A1:D1")


# Insert Empty Row & Col
ws.insert_rows(7)  # after row 6, in other words at position 7

ws.insert_cols(2)  # in position 2, in other words at B

# Delete Rows
ws.delete_rows(7)

ws.delete_cols(2)

"""


def create_workbook(workbook_name, sheets, output_filepath):
    pass


if __name__ == "__main__":
    ...
